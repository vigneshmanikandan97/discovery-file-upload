from openpyxl import  load_workbook, Workbook
from pathlib import Path
import json

def find_value_range(ws, row, col):
    """
    Finds the row limit for each not 'None' column
    returns the boundary up to which the row iteration can continue
    """
    while ws.cell(row = row, column = col).value != None:
        row += 1
    return row - 1

def xl_iterator(ws, ROWS, COLS):
    """
    Iterates the excel file and parses the content into a specific JSON format
    Sample JSON:
        ais_json = {
            serial_number: {
                "questions": [<list of parsed questions>],
                "sources": [<list of parsed source per question>],
                "relevance": [<list of parsed relevance per question>]
            }
        }
    """
    parsed_xl_json = {}
    maxr = init_row = count = row_bound = 0
    col_index = 1
    row_index = 2
    flag = False
    ROWS += 1
    xl_dict = {}

    while col_index <= COLS:    
        
        while row_index <= ROWS:
            cell_value = ws.cell(row = row_index, column = col_index).value
           
            if cell_value != None:
                
                if col_index == 1:
                    flag = False
                    count += 1
                    init_row = maxr = row_index
                    row_bound = max(row_bound, maxr)
                    xl_dict = {
                        count : {
                            "questions": [],
                            "sources": [],
                            "relevance": []
                        }
                    }

                if col_index == 2:
                    maxr = find_value_range(ws, row_index, col_index)
                    row_bound = max(row_bound, maxr)
                    question_generator = ws.iter_cols(min_col = col_index, max_col = col_index, min_row = init_row, max_row = maxr, values_only = True)
                    row_index = maxr
                    
                    for questions in question_generator:
                        for question in questions:
                            xl_dict[count]['questions'].append(question)
                
                if col_index == 3:
                    maxr = find_value_range(ws, row_index, col_index)
                    row_bound = max(row_bound, maxr)
                    source_generator = ws.iter_cols(min_col = col_index, max_col = col_index, min_row = init_row, max_row = maxr, values_only = True)
                    row_index = maxr
                    
                    for sources in source_generator:
                        for source in sources:
                            xl_dict[count]['sources'].append(source)
                
                if col_index == 4:
                    maxr = find_value_range(ws, row_index, col_index)
                    row_bound = max(row_bound, maxr)
                    relevance_generator = ws.iter_cols(min_col = col_index, max_col = col_index, min_row = init_row, max_row = maxr, values_only = True)
                    row_index = row_bound + 1
                    flag = True

                    for relevances in relevance_generator:
                        for relevance in relevances:
                            xl_dict[count]['relevance'].append(relevance)

                # Update dictionary
                parsed_xl_json[count] = xl_dict[count].copy()
                
                if row_index > row_bound:
                    break
                else:
                    row_index += 1

            else:
                """ 
                if current cell has no value (is None)
                    1. make row index as first index where data was found
                    2. increment column index
                    3. come out of row iteration
                """
                
                if row_index == (row_bound + 1) and flag == True:
                    init_row = row_bound + 2
                    
                    if col_index == 1 and row_index == ROWS:
                        # print('\n---End of Excel file---\n')
                        return parsed_xl_json
                    
                    col_index = 0

                row_index = init_row
                
                break
        
        if col_index == COLS and row_index <= ROWS:
            col_index = 0
        col_index += 1

if __name__ == "__main__":
    # Load workbook and worksheet
    wb = load_workbook(Path('/discovery-file-upload/Assets/mock_data.xlsx'))
    ws = wb['Sheet1']

    # Get total records in the worksheet
    dimensions = ws.calculate_dimension()
    ROWS = int(dimensions.split('A1:D')[1])
    # since format has only 4 columns by default
    COLS = 4
    
    resJSON = xl_iterator(ws, ROWS, COLS)
    print(json.dumps(resJSON, indent=4))